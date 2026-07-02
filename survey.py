from pathlib import Path
from io import BytesIO
import csv

import streamlit as st
from openpyxl import Workbook, load_workbook

FILE_PATH = Path(__file__).with_name("survey_results.xlsx")
CSV_PATH = Path(__file__).with_name("survey_results.csv")
HEADERS = ["名前", "年齢", "性別", "症状", "薬"]


def get_or_create_sheet():
    if FILE_PATH.exists():
        workbook = load_workbook(FILE_PATH)
        if "アンケート結果" in workbook.sheetnames:
            sheet = workbook["アンケート結果"]
        else:
            sheet = workbook.create_sheet("アンケート結果")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "アンケート結果"

    if sheet.max_row == 0 or sheet.cell(row=1, column=1).value is None:
        sheet.append(HEADERS)
    elif list(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) != [tuple(HEADERS)]:
        sheet.insert_rows(1)
        sheet.append(HEADERS)

    return workbook, sheet


def export_sheet_to_csv(sheet) -> None:
    rows = list(sheet.iter_rows(values_only=True))
    with CSV_PATH.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(rows)


def export_workbook_to_bytes() -> bytes:
    workbook, _ = get_or_create_sheet()
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()


def save_answer(data: dict[str, str]) -> None:
    workbook, sheet = get_or_create_sheet()
    sheet.append([data["名前"], data["年齢"], data["性別"], data["症状"], data["薬"]])
    workbook.save(FILE_PATH)
    export_sheet_to_csv(sheet)


def load_answers() -> list[list[str]]:
    if not FILE_PATH.exists():
        return []

    workbook = load_workbook(FILE_PATH)
    if "アンケート結果" not in workbook.sheetnames:
        return []

    sheet = workbook["アンケート結果"]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    return rows[1:]


def delete_answer(row_index: int) -> None:
    workbook = load_workbook(FILE_PATH)
    sheet = workbook["アンケート結果"]
    sheet.delete_rows(row_index + 2, 1)
    workbook.save(FILE_PATH)
    export_sheet_to_csv(sheet)


st.set_page_config(page_title="アンケート入力", page_icon="📝", layout="centered")

st.markdown(
    """
    <style>
    .stApp { background: linear-gradient(135deg, #fef7ff 0%, #eef2ff 50%, #ecfeff 100%); }
    .block-container { padding-top: 2rem; padding-bottom: 2rem; }
    .title { color: #6d28d9; font-weight: 800; }
    .card {
        background: rgba(255,255,255,0.92);
        border: 1px solid #e9d5ff;
        border-radius: 16px;
        padding: 1rem 1.1rem;
        box-shadow: 0 10px 25px rgba(109, 40, 217, 0.12);
        margin-bottom: 0.8rem;
    }
    .stButton > button {
        border-radius: 999px;
        border: none;
        background: linear-gradient(90deg, #8b5cf6, #ec4899);
        color: white;
    }
    .stTextInput > div > div > input,
    .stNumberInput input,
    .stSelectbox > div > div > div,
    .stMultiSelect > div > div > div {
        border-radius: 10px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown('<h1 class="title">🌸 アンケート入力</h1>', unsafe_allow_html=True)
st.caption("画面上のボタンやチェックボックスをクリックして入力してください。")

st.markdown("---")

st.markdown('<div class="card">', unsafe_allow_html=True)
st.subheader("新規登録")
with st.form("survey_form"):
    name = st.text_input("名前")
    age = st.number_input("年齢", min_value=0, max_value=120, step=1)
    gender = st.radio("性 別", ["男性", "女性"], horizontal=True)
    symptoms = st.multiselect("症状", ["頭痛", "風邪", "目の疲れ", "便秘", "その他"])
    medicine = st.radio("薬", ["鎮痛剤", "解熱剤", "風邪薬", "目薬", "その他"], horizontal=True)

    submitted = st.form_submit_button("保存する", use_container_width=True)

if submitted:
    if not name:
        st.warning("名前を入力してください。")
    else:
        save_answer(
            {
                "名前": name,
                "年齢": str(age),
                "性別": gender,
                "症状": ", ".join(symptoms),
                "薬": medicine,
            }
        )
        st.success("保存しました。")
        st.info(f"Excel: {FILE_PATH}\nCSV: {CSV_PATH}")
        st.rerun()

st.markdown("</div>", unsafe_allow_html=True)

st.download_button(
    "Excel出力",
    data=export_workbook_to_bytes(),
    file_name="survey_results.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

st.markdown("---")

answers = load_answers()

if answers:
    search_word = st.text_input("検索", placeholder="名前・症状・薬で検索")
    filtered_answers = []
    for original_index, row in enumerate(answers):
        haystack = " ".join([str(value or "") for value in row]).lower()
        if search_word.lower() in haystack:
            filtered_answers.append((original_index, row))

    if filtered_answers:
        for original_index, row in filtered_answers:
            col1, col2 = st.columns([5, 1])
            with col1:
                st.markdown(
                    f"""
                    <div class="card">
                        <strong>{row[0] or '名無し'}</strong><br>
                        年齢: {row[1]}<br>
                        性別: {row[2]}<br>
                        症状: {row[3]}<br>
                        薬: {row[4]}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            with col2:
                if st.button("削除", key=f"delete_{original_index}", use_container_width=True):
                    st.session_state["pending_delete_index"] = original_index
                    st.session_state["pending_delete_name"] = row[0] or "この件"

        if st.session_state.get("pending_delete_index") is not None:
            pending_index = st.session_state["pending_delete_index"]
            pending_name = st.session_state.get("pending_delete_name", "この件")
            st.warning(f"{pending_name} を削除しますか？")
            col_a, col_b = st.columns(2)
            with col_a:
                if st.button("削除する", key="confirm_delete_yes"):
                    delete_answer(pending_index)
                    st.session_state["pending_delete_index"] = None
                    st.session_state["pending_delete_name"] = ""
                    st.rerun()
            with col_b:
                if st.button("キャンセル", key="confirm_delete_no"):
                    st.session_state["pending_delete_index"] = None
                    st.session_state["pending_delete_name"] = ""
                    st.rerun()
    else:
        st.info("検索条件に一致するデータはありません。")
else:
    st.info("まだ登録された内容はありません。")
